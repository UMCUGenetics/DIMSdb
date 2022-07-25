#!/usr/bin/env python3

"""
Usage: Import Alissa export in the form of .xsls into a MongoBD database.
       Make sure the data is in a sub folder called 'input'.
Python version: 3.7.1

Input .xlsx format
[Header] - spans row 1 (dependant on #samples)
[General Metadata] -
    - 12 columns in between intensity and z-score of samples (dependant on #samples)
[data]
    - intensity Column B until "HMDB_name" metadata (dependant on the #samples)
    - Z-score Column after "sd.ctrls" (dependant on the #samples)
"""

import os
import io
import time
import datetime
import warnings
import argparse
import sys
import logging
import logging.config
import yaml
from openpyxl import load_workbook
import sqlite3
import pyreadr


# SQLite database queries and table creators and connection
connection = sqlite3.connect("DIMSdb.db")
with connection:
    connection.execute("""
        CREATE TABLE IF NOT EXISTS patientdata( 
        id INTEGER PRIMARY KEY, zscore REAL, intensity REAL,
        run_id TEXT, patient_id TEXT, hmdb_id TEXT, 
        FOREIGN KEY(hmdb_id) REFERENCES hmdbdata(hmdb_code),
        FOREIGN KEY(run_id) REFERENCES rundata(run_name),
        FOREIGN KEY(patient_id) REFERENCES patientmetadata(monsternr),
        UNIQUE(run_id, patient_id, hmdb_id)
        );
        """)

    connection.execute("""
        CREATE TABLE IF NOT EXISTS hmdbdata(
        hmdb_code TEXT PRIMARY KEY, description TEXT, origin TEXT, fluids TEXT, tissue TEXT, disease TEXT, pathway TEXT, 
        hmdb_name TEXT, relevance TEXT, db_version TEXT,
        UNIQUE(hmdb_code, db_version)
        );
        """)
    connection.execute("""
        CREATE TABLE IF NOT EXISTS patientmetadata(
        monsternr TEXT PRIMARY KEY, fullname TEXT, date_of_birth TEXT
        );
        """)
    connection.execute("""
        CREATE TABLE IF NOT EXISTS controldata(
        id INTEGER PRIMARY KEY, zscore REAL, intensity REAL, avg_ctrls REAL, sd_ctrls REAL, 
        patient_id TEXT, run_id TEXT, hmdb_id TEXT,
        FOREIGN KEY(hmdb_id) REFERENCES hmdbdata(hmdb_code),
        FOREIGN KEY(run_id) REFERENCES rundata(run_name),
        FOREIGN KEY(patient_id) REFERENCES patientmetadata(monsternr),
        UNIQUE(run_id, patient_id, hmdb_id)
        );
        """)
    connection.execute("""
        CREATE TABLE IF NOT EXISTS rundata(
        run_name TEXT PRIMARY KEY, processed_date TEXT, matrix TEXT, upload_time TEXT
        );
        """)

INSERT_HMDBDATA = '''
    INSERT OR IGNORE INTO hmdbdata (hmdb_name, relevance, description, origin, fluids, tissue, disease, pathway, hmdb_code, db_version)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
'''
INSERT_PATIENTDATA = '''
    INSERT OR IGNORE INTO patientdata (patient_id, zscore, intensity, run_id, hmdb_id) 
    VALUES (?, ?, ?, ?, ?);
'''
INSERT_CONTROLDATA = '''
    INSERT OR IGNORE INTO controldata (patient_id, zscore, intensity, run_id, hmdb_id, avg_ctrls, sd_ctrls) 
    VALUES (?, ?, ?, ?, ?, ?, ?);
'''
INSERT_RUNDATA = '''
    INSERT OR IGNORE INTO rundata (run_name, processed_date, matrix, upload_time) 
    VALUES (?, ?, ?, ?);
'''
INSERT_PATIENTMETADATA = '''
    INSERT OR IGNORE INTO patientmetadata (monsternr, fullname, date_of_birth) 
    VALUES (?, ?, ?);
'''


def add_hmdbdata(connection, hmdbdatarow):
    with connection:
        connection.execute(INSERT_HMDBDATA, (hmdbdatarow))


def add_patientdata(connection, patientdatarow):
    with connection:
        connection.execute(INSERT_PATIENTDATA, (patientdatarow))


def add_controldata(connection, controldatarow):
    with connection:
        connection.execute(INSERT_CONTROLDATA, (controldatarow))


def add_rundata(connection, rundatarow):
    with connection:
        connection.execute(INSERT_RUNDATA, (rundatarow))


def add_patientmetadata(connection, patientmetadatarow):
    with connection:
        connection.execute(INSERT_PATIENTMETADATA, (patientmetadatarow))


# configuration for the logging file. This file logs the import of excel files into MongoDB and errors
# with open(os.path.dirname(__file__) + "/logging_config.yml", 'r') as configfile:
with open("/Users/fdegruyt/github_scripts/DIMSdb_main/DIMSdb/extra/logging_config.yml", 'r') as configfile:
    logging.config.dictConfig(yaml.safe_load(configfile))

logger = logging.getLogger(__name__)
logger.debug("File upload script 'import_data.py' started")
start_time = time.time()
upload_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')


def argparser():
    # function to enable commandline help functions. Makes use of the argparse installation.
    # also makes sure that required parameters are used, with basic file extension checks

    # Override of the error function in ArgumentParser from argparse
    # To display the original error message + help page in the terminal, when the wrong or no arguments are given
    class DefaultHelpParser(argparse.ArgumentParser):
        def error(self, message):
            sys.stderr.write('\nerror: %s\n' % message)
            self.print_help()
            logger.debug('error: %s' % message)
            sys.exit(1)  # exit program

    # Override of adding an extra action class in Action from argparse
    # To add an extra action that can check for file extensions and display the help page if wrong file extension
    def CheckExt(choices):
        class Act(argparse.Action):
            def __call__(self, parser, namespace, fname, option_string=None):
                for f in fname:
                    ext = os.path.splitext(f)[1][1:]
                    if ext not in choices:
                        option_string = '({})'.format(option_string) if option_string else ''
                        parser.error("file doesn't end with {}{}".format(choices, option_string))  # display error message
                    else:
                        setattr(namespace, self.dest, fname)
        return Act

    # required arguments and optional arguments for the terminal. A help page is automatically generated.
    parser = DefaultHelpParser(
        description=(
            "DIMSdb import data:\n"
            "Imports the results of the DIMS pipeline after transferring this data into\n"
            "the Dwergeik UMCU server. An additional script will notice new files in this folder\n"
            "and use these as a parameter for this script. This will update the local SQL database."
        ),
        formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument(
        "-v",
        "--version",
        action='version',
        version='%(prog)s 1.0')
    # puts following arguments under 'required', rather than the default 'optional'
    required_argument = parser.add_argument_group('required arguments')
    # add other file-extensions if needed: for example: CheckExt({'xlsx', 'xls})
    required_argument.add_argument(
        "-f",
        "--file",
        metavar='',
        required=True,
        nargs='+',
        action=CheckExt({'xlsx', 'RData', 'rds'}),
        help=(".xlsx, .RData input, usage: -f /path/to/files/XXX.xlsx (or if multiple files, "
              "seperate with space: -f /path/to/files/XXX.xlsx /path/to/files/XXX2.xlsx'"))
    args = parser.parse_args()
    return args.file


def main():
    '''Main function, for extracting the data from the excelformat and parse to the database'''
    files = argparser()  # get the files from the -f parameter
    # quiet the warning when loading the first xlsx
    warnings.simplefilter("ignore")
    for file in files:
        filename = os.path.splitext(os.path.basename(file))[0]
        fileextension = os.path.splitext(os.path.basename(file))[1]
        if fileextension == ".RData":
            hmdb_loader(filename, file)
        elif fileextension == ".xlsx":
            excel_loader(filename, file)
            #patient_loader(filename, file)
        else:
            sys.exit("error: file doesn't end with '.xlsx' or '.RData")


def patient_loader(filename, file):
    print("hoi_patient_loader")


def hmdb_loader(filename, file):
    # fill hmdbdata table
    result = pyreadr.read_r(file)  # also works for Rds
    print(result.keys())  # let's check what objects we got
    df1 = result["rlvnc"]
    print(df1[:2])

    hallo = hoi

    if file.lower().endswith(".xlsx") and not file.lower().startswith("~"):

        with open(file, "rb") as f:
            in_mem_file = io.BytesIO(f.read())
        wb = load_workbook(in_mem_file)
        sheet = wb.active

        header_list = []
        for cell in sheet[1]:
            header_list.append(cell.value)
        # from all headers, extract the patient columns. control samples start with "C" and others with "P"
        res = [idx for idx in header_list if idx.lower().startswith("C".lower()) or idx.lower().startswith("P".lower())]
        total_samples = int(((len(res) - 2) / 2))  # -2 of default columnnames "plot" and "pathway", which are not samples

        # get the correct indexes for the metadata and patient columns
        startcol_intensities = 1  # first column is always "plot", second starts the data
        endcol_intensities = startcol_intensities + total_samples - 1
        startcol_meta = endcol_intensities + 1  # metadate between intensity and Zscore data (12 columns)
        endcol_meta = startcol_meta + 12 - 1  # number of metadata columns between intensity and Zscore
        startcol_zscore = endcol_meta + 1  # Zscores start after the metadata
        endcol_zscore = startcol_zscore + total_samples - 1  # is not used in this code yet. here for reference

        for i in range(2, sheet.max_row+1):
            hmdbdatarow = [cell.value for cell in sheet[i][startcol_meta+1:endcol_meta-1]]
            hmdbdatarow.append(filename)
            add_hmdbdata(connection, hmdbdatarow)

    logger.debug('finished uploading: %s' % file)
    print("## \t\tFinished in {0:.2f} minutes".format((time.time() - start_time) / 60))


def excel_loader(filename, file):
    logger.debug('start uploading: %s' % file)
    # To prevent opening a cached version of the file
    if file.lower().endswith(".xlsx") and not file.lower().startswith("~"):
        with open(file, "rb") as f:
            in_mem_file = io.BytesIO(f.read())
        wb = load_workbook(in_mem_file)
        sheet = wb.active

        header_list = []
        for cell in sheet[1]:
            header_list.append(cell.value)
        # from all headers, extract the patient columns. control samples start with "C" and others with "P"
        res = [idx for idx in header_list if idx.lower().startswith("C".lower()) or idx.lower().startswith("P".lower())]
        total_samples = int(((len(res) - 2) / 2))  # -2 of default columnnames "plot" and "pathway", which are not samples

        # get the correct indexes for the metadata and patient columns
        startcol_intensities = 1  # first column is always "plot", second starts the data
        endcol_intensities = startcol_intensities + total_samples - 1
        startcol_meta = endcol_intensities + 1  # metadate between intensity and Zscore data (12 columns)
        endcol_meta = startcol_meta + 12 - 1  # number of metadata columns between intensity and Zscore
        startcol_zscore = endcol_meta + 1  # Zscores start after the metadata
        endcol_zscore = startcol_zscore + total_samples - 1  # is not used in this code yet. here for reference

        for i in range(1, sheet.max_row+1):
            # first get patient info from intensities header (patient name is same as header)
            if i == 1:
                patients = [cell.value for cell in sheet[i][startcol_intensities:endcol_intensities+1]]
            else:
                patient_index = 0
                avg_ctrls = sheet[i][endcol_meta-1].value
                sd_ctrls = sheet[i][endcol_meta].value
                hmdb_code = sheet[i][endcol_meta-2].value
                run_name = filename
                matrix = filename.split("_")[1]
                processed_date = datetime.datetime.strptime(filename.split("_")[2], '%Y%m%d').date()

                # get intensity and zscore data for each patient
                for cell in [cell.value for cell in sheet[i][startcol_intensities:endcol_intensities+1]]:
                    patient_monsternr = patients[patient_index]
                    patient_nr = patient_monsternr.split(".")[0]
                    patient_subnr = patient_monsternr.split(".")[1]
                    intensity = cell
                    # see structure at top of this file
                    zscore = sheet[i][startcol_intensities+patient_index+total_samples+12].value

                    rundata = run_name, processed_date, matrix, upload_time
                    add_rundata(connection, rundata)

                    patientmetadata = run_name, processed_date, matrix
                    add_patientmetadata(connection, patientmetadata)

                    if patient_monsternr.lower().startswith("C".lower()):
                        controldata = [patient_monsternr, zscore, intensity, run_name, hmdb_code, avg_ctrls, sd_ctrls]
                        add_controldata(connection, controldata)
                    else:
                        patientdata = [patient_monsternr, zscore, intensity, run_name, hmdb_code]
                        add_patientdata(connection, patientdata)

                    patient_index += 1  # increment to add next patient (next column)
    logger.debug('finished uploading: %s' % file)
    # Moves file to other folder, to make uploading a new file (with same filename) possible on the UMCU server (for replacing)
    #os.replace(file, os.path.dirname(__file__) + "/verwerkt/" + os.path.basename(file))
    print("## \t\tFinished in {0:.2f} minutes".format((time.time() - start_time) / 60))


main()
